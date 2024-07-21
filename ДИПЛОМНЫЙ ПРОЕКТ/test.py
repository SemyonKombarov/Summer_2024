""" Модуль предназначен для парсинга исходных данных от службы ГиР и формирования профилей покустовой добычи """
import re
import datetime
import pandas as pd
from openpyxl import load_workbook
start = datetime.datetime.now()
wb = load_workbook("/Summer_2024/Summer_2024/ДИПЛОМНЫЙ ПРОЕКТ/test_id_3.xlsx", data_only=True)
ws = wb.active
dct = {}  # Словарь который ляжет в основу датафрейма
print('Начало выполнение процедуры парсинга')
sn = wb.sheetnames
for j in sn:
    # Цикл для пробежки по листам книги
    if len(j) <= 3:
        wb.active = wb[j]
        ws = wb.active
        eo = j
        for i in range(ws.max_row):
            # Определение координаты и типа параметра на листе, отправная точка для формирования датафрейма
            if type(ws.cell(row=i + 1, column=1).value) == str:
                row_start = ws.cell(row=i + 1, column=1).row
                column_start = ws.cell(row=i + 1, column=1).column
                x = re.findall(r"(\S+),", ws.cell(row=i + 1, column=1).value)
                # Костыль
                parametr = 0
                for kostil in x:
                    parametr = kostil
                # Определение имен скважин
                for m in range(ws.max_column):
                    if ws.cell(row=i + 2, column=m + 1).value != None and ws.cell(row=i + 2,column=m + 1).value != "Dates":
                        y = re.findall(r"\d\d\d\d\d-\d", ws.cell(row=i + 2, column=m + 1).value)
                        # Костыль
                        well_name = 0
                        for kostil in y:
                            well_name = kostil
                        # Разделение имени скважины на составные чатси - ЛУ, КГС, номер и ЭО
                        ly = str(well_name)[0]
                        kgs = str(well_name)[1:3]
                        well_number = str(well_name)[3:5]
                        eks_ob = str(well_name)[-1]
                        # Определение конкретной даты
                        for mm in range(ws.max_row):
                            if type(ws.cell(row=mm + 2, column=1).value) == int and int(
                                    str(ws.cell(row=mm + 2, column=2).value)[:4]) < 2150 and ws.cell(row=mm + 2,column=1).value != 0:
                                date = str(ws.cell(row=mm + 2, column=2).value)[:4]
                                dct[(parametr, (ly, kgs, well_number, eks_ob), date)] = dct.get((parametr, (ly, kgs, well_number, eks_ob), date),
                                    ws.cell(row=i + mm, column=m + 1).value)
                                # dct[(eo,parametr,well_name, ly, kgs,well_number,eks_ob,date)]=dct.get((eo,parametr,well_name, ly, kgs,well_number,eks_ob,date),ws.cell(row = i + mm, column = m + 1).coordinate)

finish = datetime.datetime.now()
print(f'Время парсинга составило: {finish - start}')

# Запись результатов парсинга в файл, формирование ИД для датафреймов
lst = []
lst_pd = []
file_output_lite = open('output_lite.txt', 'w')
file_output_full = open('output_full.txt', 'w')
for k, v in dct.items():
    file_output_lite.writelines(f'{k}, {v}\n')
    file_output_full.writelines(f'{[k[0], k[1][0], k[1][1], k[1][2], k[1][3], k[2], v]}\n')
    lst.append([k[0], k[1], k[2], v])
    lst_pd.append([k[0], k[1][0], k[1][1], k[1][2], k[1][3], k[2], v])
file_output_lite.close()
file_output_full.close()

df = pd.DataFrame(lst, columns=['Параметр', 'Скважина', 'Год', 'Значение']) # Оптимизированная версия датафрейма
df_pd = pd.DataFrame(lst_pd,columns=['Параметр','Номер ЛУ','Номер КГС','Номер скважины','Номер ЭО','Год','Значение'])# Полная версия

#Группировка и фильтрация по датафреймам
wellpads_sum_gas_rate = df_pd[df_pd.Параметр == "rate_gas"].groupby(['Параметр',"Номер КГС",'Год']).agg({"Значение":'sum'}).rename(columns = {"Значение":"rate_gas"})
wellpads_sum_gas_rate.to_csv('wellpads_sum_gas_rate.txt', sep='\t',index =True)
wellpads_min_thp = df_pd[df_pd.Параметр == "pressure_thp"].replace(0,None).groupby(['Параметр',"Номер КГС",'Год']).agg({"Значение":'min'}).rename(columns = {"Значение":"pressure_thp"})
wellpads_min_thp.to_csv('wellpads_min_thp.txt', sep='\t',index =True)
wellpads_sum_condensate_rate = df_pd[df_pd.Параметр == "rate_cond"].groupby(['Параметр',"Номер КГС",'Год']).agg({"Значение":'sum'}).rename(columns = {"Значение":"rate_cond"})
wellpads_sum_condensate_rate.to_csv('wellpads_sum_condensate_rate.txt', sep='\t',index =True)
wellpads_sum_water_rate = df_pd[df_pd.Параметр == "rate_water"].groupby(['Параметр',"Номер КГС",'Год']).agg({"Значение":'sum'}).rename(columns = {"Значение":"rate_water"})
wellpads_sum_water_rate.to_csv('wellpads_sum_water_rate.txt', sep='\t',index =True)

ly_sum_gas_rate = (df_pd[df_pd.Параметр == "rate_gas"].groupby(['Год']).agg({"Значение":'sum'}).rename(columns = {"Значение":"ly_rate_gas"})).max().reset_index()
ly_sum_cond_rate = (df_pd[df_pd.Параметр == "rate_cond"].groupby(['Год']).agg({"Значение":'sum'}).rename(columns = {"Значение":"ly_rate_cond"})).max().reset_index()

#Merge для группы датафреймов
data = pd.merge(wellpads_sum_gas_rate, wellpads_sum_condensate_rate, how= "outer", on=["Год","Номер КГС"])
data = pd.merge(data,wellpads_sum_water_rate, how= "outer", on=["Год","Номер КГС"])
data = pd.merge(data,wellpads_min_thp, how= "outer", on=["Год","Номер КГС"])
data.to_csv('data.txt', sep='\t',index =True)

data.to_excel('data.xlsx',index = True,merge_cells=False)

total_wells = set()
for i in lst:
    total_wells.add(i[1])

total_wellpads = set()
total_eo = set()
for i in total_wells:
    total_wellpads.add(i[1])
    total_eo.add(i[3])
print()
print('КРАТКАЯ АНАЛИТИКА ПО ЛИЦЕНЗИОННОМУ УЧАСТКУ')
print(f' Общее количество эксплуатационых скважин на ЛУ - {len(total_wells)} шт')
print(f' Общее количество КГС на ЛУ - {len(total_wellpads)} шт')
print(f' Общее количество ЭО на ЛУ - {len(total_eo)} шт')
print(f' Пиковое значение добычи СОГ всего ЛУ ~ {ly_sum_gas_rate[0][0]*347/10**9:1.2f} млрд м3/год')
print(f' Пиковое значение добычи СГК всего ЛУ ~ {ly_sum_cond_rate[0][0]*347/10**6:1.2f} млн т/год')
# print(f' Пиковое значение добычи СОГ ЭО {min(total_eo)}-4 ~ {ly_sum_gas_rate[0][0]*347/10**9:1.2f} млрд м3/год')
# print(f' Пиковое значение добычи СОГ ЭО 5-{max(total_eo)} ~ {ly_sum_gas_rate[0][0]*347/10**9:1.2f} млрд м3/год')

print()
print('КРАТКАЯ АНАЛИТИКА ПО КГС')
count = 0
total_count = 0
for j in sorted(list(total_eo)):
    for i in total_wells:
        if i[3] == j:
            count += 1
            total_count += 1
    print(f' Общее число скважин на ЭО {j} - {count} шт')
    count = 0

for j in sorted(list(total_wellpads)):
    x = df_pd[df_pd["Номер КГС"] == j]
    for i in x:
        if i == "Номер скважины":
            y = x[i].unique()
            print(f" Общее количество скважин на КГС {j} - {len(y)} штук")




