import openpyxl
from openpyxl import Workbook
# wb = Workbook()
# wb.save("excel.xlsx")
wb = openpyxl.load_workbook("excel.xlsx")
# print(wb.sheetnames)
ws = wb.active
print("Входные данные")
for i in range(ws.max_row):
    for j in range(ws.max_column):
        print(i+1,j+1,ws.cell(row = i+1, column = j + 1).value)
wb.create_sheet("New_sheet")
# print(wb.sheetnames)
wb.active = wb["New_sheet"]
ws1 = wb.active
# print(wb.active)
dct = {}
lst = []
for i in range(ws.max_row):
    # print(i, "i")
    for j in range(ws.max_column):
        # print(j, "j")
        # print(i + 1, j + 1, ws1.cell(row=i + 1, column=j + 1).value)
        ws1.cell(row=i + 1, column=j + 1).value = ws.cell(row=i + 1, column=j + 1).value
        dct[ws.cell(row=i + 1, column=1).value] = dct.get(ws.cell(row=i + 1, column=1).value,ws.cell(row=i+1, column=2).value)
        # print(dct)

    lst.append(ws1.cell(row=i + 1, column=2).value)
# print(dct)
sort_dct = sorted(dct.items(), key =  lambda x: (-x[1],x[0]))
# print(sort_dct[0][0])
# print(sum(lst))
# print(wb.active)
print("Вывод по результатам сортировки")
for i in range(ws.max_row):
    for j in range(ws.max_column):
        ws1.cell(row=i + 1, column=j + 1).value = sort_dct[i][j]
        print(i + 1, j + 1, ws1.cell(row=i + 1, column=j + 1).value)
ws1.cell(row=ws1.max_row+1, column=1).value = "Итого:"
ws1.cell(row=ws1.max_row, column=2).value = sum(lst)

wb.save("excel.xlsx")