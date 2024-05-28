import openpyxl
from openpyxl import Workbook
from statistics import mean,median
# wb = Workbook()
# wb.save("excel_9_3.xlsx")

wb = openpyxl.load_workbook("excel_9_3.xlsx")
ws = wb.active

lst = []
for i in range(ws.max_row):
    lst.append(ws.cell(row=i + 1, column=2).value)
maxi = max(lst)
mini = min(lst)
summ = sum(lst)
sred = mean(lst)
mediana = median(lst)
# print(maxi,mini,summ,sred,int(mediana))
wb.create_sheet("statistics")
wb.active = wb["statistics"]
ws = wb.active
ws.cell(row=1, column=1).value = "Максимальное значение"
ws.cell(row=2, column=1).value = "Минимальное значение"
ws.cell(row=3, column=1).value = "Сумма"
ws.cell(row=4, column=1).value = "Среднее арифметическое"
ws.cell(row=5, column=1).value = "Медиана"
ws.cell(row=1, column=2).value = maxi
ws.cell(row=2, column=2).value = mini
ws.cell(row=3, column=2).value = summ
ws.cell(row=4, column=2).value = sred
ws.cell(row=5, column=2).value = mediana
wb.save("excel_9_3.xlsx")
print("Программа выполнена!")